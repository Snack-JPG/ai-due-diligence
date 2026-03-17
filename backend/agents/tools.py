from __future__ import annotations

import ast
import math
import re
from pathlib import Path
from typing import Any

import requests
from langchain.tools import tool
from openpyxl import load_workbook

from backend.config import settings
from backend.database import db
from backend.retrieval import retrieval_service


def _safe_eval(expression: str) -> float:
    allowed_nodes = {
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Num,
        ast.Constant,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Load,
        ast.Mod,
    }
    node = ast.parse(expression, mode="eval")
    for subnode in ast.walk(node):
        if type(subnode) not in allowed_nodes:
            raise ValueError("Unsupported expression")
    return float(eval(compile(node, "<calculator>", "eval"), {"__builtins__": {}}, {"math": math}))


def build_rag_tool(analysis_id: str, preferred_types: list[str] | None = None):
    @tool("rag_retrieval_tool")
    def rag_retrieval_tool(query: str) -> dict[str, Any]:
        """Retrieve relevant uploaded-document evidence for the analysis."""
        chunks = retrieval_service.search(analysis_id, query, preferred_types=preferred_types)
        return {
            "query": query,
            "chunks": chunks,
            "rendered_context": retrieval_service.format_chunks(chunks),
        }

    return rag_retrieval_tool


@tool("calculator_tool")
def calculator_tool(expression: str) -> dict[str, Any]:
    """Evaluate a numeric expression and return the result."""
    result = _safe_eval(expression)
    return {"expression": expression, "result": result}


def build_table_parser_tool(analysis_id: str):
    @tool("table_parser_tool")
    def table_parser_tool(filename: str) -> dict[str, Any]:
        """Extract rows from a spreadsheet by uploaded filename."""
        matches = [doc for doc in db.get_documents(analysis_id) if doc["filename"] == filename]
        if not matches:
            raise ValueError(f"No uploaded document named {filename}")
        path = Path(matches[0]["storage_path"])
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(filename=str(path), data_only=True)
            sheets = {}
            for sheet in workbook.worksheets:
                rows = []
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    rows.append({"row": idx, "values": ["" if value is None else str(value) for value in row]})
                sheets[sheet.title] = rows
            return {"filename": filename, "sheets": sheets}
        if path.suffix.lower() == ".csv":
            content = path.read_text(encoding="utf-8", errors="ignore").splitlines()
            rows = [line.split(",") for line in content]
            return {"filename": filename, "rows": rows}
        raise ValueError("Table parser supports XLSX and CSV files only")

    return table_parser_tool


@tool("keyword_extraction_tool")
def keyword_extraction_tool(text: str) -> dict[str, Any]:
    """Extract diligence-relevant clause keywords from a body of text."""
    patterns = [
        "indemn",
        "liability",
        "termination",
        "exclusive",
        "governing law",
        "jurisdiction",
        "assignment",
        "privacy",
        "compliance",
    ]
    matches = []
    lower_text = text.lower()
    for pattern in patterns:
        if pattern in lower_text:
            for snippet in re.findall(rf".{{0,60}}{re.escape(pattern)}.{{0,120}}", lower_text):
                matches.append({"keyword": pattern, "snippet": snippet.strip()})
    return {"matches": matches}


@tool("web_search_tool")
def web_search_tool(query: str) -> dict[str, Any]:
    """Search the web for recent market evidence."""
    if settings.tavily_api_key:
        response = requests.post(
            "https://api.tavily.com/search",
            json={"api_key": settings.tavily_api_key, "query": query, "max_results": 5},
            timeout=20,
        )
        response.raise_for_status()
        payload = response.json()
        return {"query": query, "results": payload.get("results", [])}

    response = requests.get("https://duckduckgo.com/html/", params={"q": query}, timeout=20)
    response.raise_for_status()
    html = response.text
    results = []
    for match in re.finditer(
        r'nofollow" class="result__a" href="(?P<url>[^"]+)">(?P<title>.*?)</a>.*?result__snippet">(?P<snippet>.*?)</a>',
        html,
        flags=re.DOTALL,
    ):
        results.append(
            {
                "title": re.sub("<.*?>", "", match.group("title")),
                "url": match.group("url"),
                "content": re.sub("<.*?>", "", match.group("snippet")),
            }
        )
        if len(results) >= 5:
            break
    return {"query": query, "results": results}
